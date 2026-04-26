#!/usr/bin/env python3
"""
price_history.py -- Price History Tracker (#5)
Plugin for Keith's PokeBS tracker system.

Logs product prices to a SQLite database once per hour.
Provides query functions used by the dashboard price history page.
Auto-purges records older than 90 days to keep the DB lean.

Storage:
  data/price_history.db  -- SQLite database (~1-3MB after 90 days)

Display:
  - Rolling 30 days shown on dashboard price history page
  - Full 90 days available for export

Export:
  python plugins/price_history.py --export
  Writes data/price_history_export.xlsx with full 90-day dataset

Schema:
  price_records (
    id          INTEGER PRIMARY KEY,
    recorded_at TEXT,       -- ISO timestamp
    name        TEXT,       -- product name
    retailer    TEXT,       -- target/walmart/bestbuy/pokemoncenter
    url         TEXT,       -- product URL
    price       REAL,       -- parsed price or NULL if N/A
    price_str   TEXT,       -- raw price string e.g. "$49.99"
    in_stock    INTEGER,    -- 1 or 0
    msrp        REAL,       -- MSRP at time of recording
    pct_of_msrp REAL        -- price/msrp * 100 or NULL
  )
"""

import os
import sys
import sqlite3
import logging
from datetime import datetime, timedelta

# -- Path resolution ----------------------------------------------------------
_here = os.path.dirname(os.path.abspath(__file__))
_root = os.path.dirname(_here) if os.path.basename(_here) == "plugins" else _here
if _root not in sys.path:
    sys.path.insert(0, _root)
if _here not in sys.path:
    sys.path.insert(0, _here)
# ----------------------------------------------------------------------------

from shared import DATA_DIR, get_msrp, parse_price

log = logging.getLogger(__name__)

DB_PATH        = os.path.join(DATA_DIR, "price_history.db")
RETAIN_DAYS    = 90   # keep records for this many days
DISPLAY_DAYS   = 30   # dashboard shows this rolling window
LOG_INTERVAL   = 60   # minutes between DB writes


# -- Database setup -----------------------------------------------------------

def get_db() -> sqlite3.Connection:
    """Return a connection to the price history database."""
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")   # safer concurrent writes
    conn.execute("PRAGMA synchronous=NORMAL") # faster writes, still safe
    return conn


def init_db() -> None:
    """Create table and indexes if they don't exist yet."""
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS price_records (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                recorded_at TEXT    NOT NULL,
                name        TEXT    NOT NULL,
                retailer    TEXT    NOT NULL,
                url         TEXT    NOT NULL,
                price       REAL,
                price_str   TEXT,
                in_stock    INTEGER NOT NULL DEFAULT 0,
                msrp        REAL,
                pct_of_msrp REAL
            )
        """)
        # Index for fast dashboard queries
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_name_retailer_time
            ON price_records (name, retailer, recorded_at DESC)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_recorded_at
            ON price_records (recorded_at DESC)
        """)
        conn.commit()
    log.debug(f"[price_history] DB initialised at {DB_PATH}")


# -- Core logging -------------------------------------------------------------

def log_prices(products: list) -> int:
    """
    Write one price record per product to the database.
    Reads current prices from status_snapshot.json via the products list.
    Returns number of records written.
    """
    import json

    # Load latest snapshot for current prices + stock status
    snapshot_path = os.path.join(DATA_DIR, "status_snapshot.json")
    try:
        with open(snapshot_path, encoding="utf-8") as f:
            snapshot = {p["url"]: p for p in json.load(f)}
    except Exception as e:
        log.warning(f"[price_history] Could not read snapshot: {e}")
        return 0

    now     = datetime.now().isoformat()
    records = []

    for product in products:
        url  = product.get("url", "")
        name = product.get("name", "")
        retailer = product.get("retailer", "")

        snap = snapshot.get(url, {})
        price_str = snap.get("price", "N/A")
        in_stock  = 1 if snap.get("in_stock", False) else 0
        price_num = parse_price(price_str)
        msrp      = get_msrp(name, retailer)
        pct       = round((price_num / msrp) * 100, 1) if (price_num and msrp) else None

        records.append((
            now, name, retailer, url,
            price_num, price_str, in_stock,
            msrp, pct
        ))

    if not records:
        return 0

    with get_db() as conn:
        conn.executemany("""
            INSERT INTO price_records
                (recorded_at, name, retailer, url, price, price_str,
                 in_stock, msrp, pct_of_msrp)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, records)
        conn.commit()

    log.info(f"[price_history] Logged {len(records)} price records")
    return len(records)


def purge_old_records() -> int:
    """Delete records older than RETAIN_DAYS. Returns count deleted."""
    cutoff = (datetime.now() - timedelta(days=RETAIN_DAYS)).isoformat()
    with get_db() as conn:
        cur = conn.execute(
            "DELETE FROM price_records WHERE recorded_at < ?", (cutoff,)
        )
        conn.commit()
        deleted = cur.rowcount
    if deleted:
        log.info(f"[price_history] Purged {deleted} records older than {RETAIN_DAYS} days")
    return deleted


# -- Query functions (used by dashboard JSON endpoint) -----------------------

def get_price_history(
    name: str = None,
    retailer: str = None,
    days: int = DISPLAY_DAYS,
    url: str = None,
) -> list[dict]:
    """
    Return price history records for the dashboard.
    Filter by name, retailer, url, or days window.
    """
    cutoff = (datetime.now() - timedelta(days=days)).isoformat()
    conditions = ["recorded_at >= ?"]
    params     = [cutoff]

    if url:
        conditions.append("url = ?")
        params.append(url)
    elif name:
        conditions.append("name LIKE ?")
        params.append(f"%{name}%")

    if retailer and not url:
        conditions.append("retailer = ?")
        params.append(retailer)

    where = " AND ".join(conditions)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT recorded_at, name, retailer, url,
                   price, price_str, in_stock, msrp, pct_of_msrp
            FROM price_records
            WHERE {where}
            ORDER BY recorded_at ASC
        """, params).fetchall()

    return [dict(r) for r in rows]


def get_price_summary(days: int = DISPLAY_DAYS) -> list[dict]:
    """
    Return a summary row per product showing:
    min, max, avg price + latest price + lowest point date.
    Used by the price history dashboard page.
    """
    cutoff = (datetime.now() - timedelta(days=days)).isoformat()
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                pr.name,
                pr.retailer,
                pr.url,
                MAX(pr.msrp)                                        AS msrp,
                MIN(pr.price)                                        AS min_price,
                MAX(pr.price)                                        AS max_price,
                ROUND(AVG(pr.price), 2)                             AS avg_price,
                SUM(CASE WHEN pr.in_stock = 1 THEN 1 ELSE 0 END)   AS times_in_stock,
                COUNT(*)                                             AS record_count,
                MAX(pr.recorded_at)                                  AS last_seen,
                (SELECT p2.price_str FROM price_records p2
                 WHERE p2.url = pr.url
                 ORDER BY p2.recorded_at DESC LIMIT 1)              AS latest_price_str,
                (SELECT p2.in_stock FROM price_records p2
                 WHERE p2.url = pr.url
                 ORDER BY p2.recorded_at DESC LIMIT 1)              AS latest_in_stock,
                (SELECT p2.recorded_at FROM price_records p2
                 WHERE p2.url = pr.url
                 ORDER BY p2.price ASC, p2.recorded_at ASC LIMIT 1) AS min_price_at
            FROM price_records pr
            WHERE pr.recorded_at >= ? AND pr.price IS NOT NULL
            GROUP BY pr.url
            ORDER BY pr.name, pr.retailer
        """, (cutoff,)).fetchall()

    return [dict(r) for r in rows]


def get_sparkline_data(url: str, days: int = 30) -> list[dict]:
    """
    Return hourly price points for a single product URL.
    Used to draw sparklines on dashboard cards.
    Downsamples to max 168 points (one per hour for 7 days)
    or the actual hourly records for longer windows.
    """
    cutoff = (datetime.now() - timedelta(days=days)).isoformat()
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                strftime('%Y-%m-%dT%H:00:00', recorded_at) AS hour,
                AVG(price)  AS price,
                MAX(in_stock) AS in_stock
            FROM price_records
            WHERE url = ?
              AND recorded_at >= ?
              AND price IS NOT NULL
            GROUP BY hour
            ORDER BY hour ASC
        """, (url, cutoff)).fetchall()

    return [dict(r) for r in rows]


def get_db_stats() -> dict:
    """Return database statistics for the help/diagnostics page."""
    with get_db() as conn:
        total = conn.execute("SELECT COUNT(*) FROM price_records").fetchone()[0]
        oldest = conn.execute(
            "SELECT MIN(recorded_at) FROM price_records"
        ).fetchone()[0]
        newest = conn.execute(
            "SELECT MAX(recorded_at) FROM price_records"
        ).fetchone()[0]
        products = conn.execute(
            "SELECT COUNT(DISTINCT url) FROM price_records"
        ).fetchone()[0]

    db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    return {
        "total_records": total,
        "oldest_record": oldest,
        "newest_record": newest,
        "unique_products": products,
        "db_size_kb": round(db_size / 1024, 1),
        "retain_days": RETAIN_DAYS,
        "display_days": DISPLAY_DAYS,
    }


# -- JSON export for dashboard ------------------------------------------------

def export_summary_json() -> bool:
    """
    Write price_summary.json to DATA_DIR for the dashboard to read.
    Called hourly alongside log_prices().
    """
    import json

    try:
        summary = get_price_summary(days=DISPLAY_DAYS)
        sparklines = {}
        # Only generate sparklines for products with enough data
        for row in summary:
            sparklines[row["url"]] = get_sparkline_data(row["url"], days=7)

        output = {
            "last_updated":  datetime.now().isoformat(),
            "display_days":  DISPLAY_DAYS,
            "retain_days":   RETAIN_DAYS,
            "summary":       summary,
            "sparklines":    sparklines,
            "stats":         get_db_stats(),
        }

        path = os.path.join(DATA_DIR, "price_summary.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        log.debug(f"[price_history] price_summary.json updated ({len(summary)} products)")
        return True
    except Exception as e:
        log.warning(f"[price_history] JSON export error: {e}")
        return False


# -- Excel export -------------------------------------------------------------

def export_excel() -> str:
    """
    Export full 90-day price history to Excel.
    Returns path to the written file.
    Requires openpyxl (pip install openpyxl).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl not installed. Run: pip install openpyxl"
        )

    out_path = os.path.join(DATA_DIR, "price_history_export.xlsx")
    wb = openpyxl.Workbook()

    # -- Sheet 1: Summary ----------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary (30 days)"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1a1a2e")
    center_align = Alignment(horizontal="center")

    sum_headers = [
        "Product", "Retailer", "MSRP", "Latest Price",
        "Min Price", "Max Price", "Avg Price",
        "Times In Stock", "Records", "Last Seen"
    ]
    ws_sum.append(sum_headers)
    for col, _ in enumerate(sum_headers, 1):
        cell = ws_sum.cell(1, col)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center_align

    summary = get_price_summary(days=DISPLAY_DAYS)
    for row in summary:
        ws_sum.append([
            row["name"],
            row["retailer"],
            f"${row['msrp']:.2f}"    if row["msrp"]      else "N/A",
            row["latest_price_str"]  or "N/A",
            f"${row['min_price']:.2f}" if row["min_price"] else "N/A",
            f"${row['max_price']:.2f}" if row["max_price"] else "N/A",
            f"${row['avg_price']:.2f}" if row["avg_price"] else "N/A",
            row["times_in_stock"],
            row["record_count"],
            row["last_seen"][:16]    if row["last_seen"]  else "N/A",
        ])

    # Auto-width columns
    for col in ws_sum.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # -- Sheet 2: Full raw data (90 days) ------------------------------------
    ws_raw = wb.create_sheet("Raw Data (90 days)")
    raw_headers = [
        "Recorded At", "Product", "Retailer", "Price",
        "Price String", "In Stock", "MSRP", "% of MSRP", "URL"
    ]
    ws_raw.append(raw_headers)
    for col, _ in enumerate(raw_headers, 1):
        cell = ws_raw.cell(1, col)
        cell.font  = header_font
        cell.fill  = header_fill

    records = get_price_history(days=RETAIN_DAYS)
    for row in records:
        ws_raw.append([
            row["recorded_at"][:16],
            row["name"],
            row["retailer"],
            row["price"],
            row["price_str"],
            "Yes" if row["in_stock"] else "No",
            row["msrp"],
            row["pct_of_msrp"],
            row["url"],
        ])

    # Auto-width key columns
    for i in [1, 2, 3, 4, 5, 6]:
        col_letter = get_column_letter(i)
        max_len = max(
            len(str(ws_raw.cell(r, i).value or ""))
            for r in range(1, min(ws_raw.max_row + 1, 100))
        )
        ws_raw.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # -- Sheet 3: By retailer pivot ------------------------------------------
    ws_ret = wb.create_sheet("By Retailer")
    retailers = ["target", "walmart", "bestbuy", "pokemoncenter"]
    ws_ret.append(["Retailer", "Products Tracked", "Avg Price vs MSRP", "Times In Stock (30d)"])
    for cell in ws_ret[1]:
        cell.font = header_font
        cell.fill = header_fill

    for ret in retailers:
        ret_rows = [r for r in summary if r["retailer"] == ret]
        if not ret_rows:
            continue
        avg_pct = round(sum(
            (r["avg_price"] / r["msrp"] * 100)
            for r in ret_rows
            if r["avg_price"] and r["msrp"]
        ) / max(len(ret_rows), 1), 1)
        total_in_stock = sum(r["times_in_stock"] for r in ret_rows)
        ws_ret.append([
            ret.title(),
            len(ret_rows),
            f"{avg_pct}% of MSRP",
            total_in_stock,
        ])

    wb.save(out_path)
    log.info(f"[price_history] Excel exported to {out_path}")
    return out_path


# -- CSV export ---------------------------------------------------------------

def export_csv() -> dict:
    """
    Export price history to two CSV files — no dependencies required.
    Returns dict with paths to both files.

    Files written to data/:
      price_summary_30d.csv   -- one row per product, 30-day summary
      price_history_90d.csv   -- full raw hourly records, 90 days
    """
    import csv

    paths = {}

    # -- Summary CSV (30 days) ------------------------------------------------
    summary_path = os.path.join(DATA_DIR, "price_summary_30d.csv")
    summary      = get_price_summary(days=DISPLAY_DAYS)

    with open(summary_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Product", "Retailer", "MSRP", "Latest Price",
            "Min Price (30d)", "Max Price (30d)", "Avg Price (30d)",
            "Times In Stock (30d)", "Records (30d)", "Last Seen", "URL"
        ])
        for row in summary:
            writer.writerow([
                row["name"],
                row["retailer"],
                f"{row['msrp']:.2f}"      if row["msrp"]      else "",
                row["latest_price_str"]   or "",
                f"{row['min_price']:.2f}" if row["min_price"] else "",
                f"{row['max_price']:.2f}" if row["max_price"] else "",
                f"{row['avg_price']:.2f}" if row["avg_price"] else "",
                row["times_in_stock"],
                row["record_count"],
                row["last_seen"][:16]     if row["last_seen"]  else "",
                row["url"],
            ])

    paths["summary"] = summary_path
    log.info(f"[price_history] Summary CSV written to {summary_path}")

    # -- Raw data CSV (90 days) -----------------------------------------------
    raw_path = os.path.join(DATA_DIR, "price_history_90d.csv")
    records  = get_price_history(days=RETAIN_DAYS)

    with open(raw_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Recorded At", "Product", "Retailer",
            "Price ($)", "Price String", "In Stock",
            "MSRP ($)", "% of MSRP", "URL"
        ])
        for row in records:
            writer.writerow([
                row["recorded_at"][:16],
                row["name"],
                row["retailer"],
                f"{row['price']:.2f}"       if row["price"]       else "",
                row["price_str"]            or "",
                "Yes" if row["in_stock"]    else "No",
                f"{row['msrp']:.2f}"        if row["msrp"]        else "",
                f"{row['pct_of_msrp']:.1f}" if row["pct_of_msrp"] else "",
                row["url"],
            ])

    paths["raw"] = raw_path
    log.info(f"[price_history] Raw CSV written to {raw_path}")
    return paths



class PriceHistoryTracker:
    """Plugin wrapper -- registered via plugins.py PriceHistory_Plugin."""

    def __init__(self, config: dict, products: list):
        self.config   = config
        self.products = products
        init_db()

    def start(self, schedule) -> None:
        # Log prices and update JSON once per hour on the hour
        schedule.every(LOG_INTERVAL).minutes.do(self._hourly_log)
        # Also run immediately so data is available right away
        self._hourly_log()
        log.info(f"[price_history] Scheduled hourly logging to {DB_PATH}")

    def _hourly_log(self) -> None:
        log_prices(self.products)
        export_summary_json()
        purge_old_records()

    def stop(self) -> None:
        log.info("[price_history] Stopped")


# -- Standalone diagnostic / export ------------------------------------------

def run_diagnostics(config: dict, products: list) -> None:
    """
    Show database stats and optionally export to Excel.
    Usage:
      python plugins/price_history.py           # show stats
      python plugins/price_history.py --export  # export to Excel
      python plugins/price_history.py --log     # force a price log now
    """
    import sys

    init_db()
    stats = get_db_stats()

    print("\n" + "=" * 55)
    print("  Price History Tracker -- Diagnostic")
    print("=" * 55)
    print(f"\n  Database:        {DB_PATH}")
    print(f"  Size:            {stats['db_size_kb']} KB")
    print(f"  Total records:   {stats['total_records']:,}")
    print(f"  Unique products: {stats['unique_products']}")
    print(f"  Oldest record:   {stats['oldest_record'] or 'none yet'}")
    print(f"  Newest record:   {stats['newest_record'] or 'none yet'}")
    print(f"  Retention:       {RETAIN_DAYS} days")
    print(f"  Display window:  {DISPLAY_DAYS} days")

    if "--log" in sys.argv:
        print("\n  Forcing price log now...")
        n = log_prices(products)
        export_summary_json()
        print(f"  Logged {n} records.")

    if "--csv" in sys.argv:
        print("\n  Exporting to CSV...")
        paths = export_csv()
        print(f"  Summary (30d): {paths['summary']}")
        print(f"  Raw data (90d): {paths['raw']}")
        print("  Open either file in Google Sheets, LibreOffice, or any text editor.")

    if "--export" in sys.argv:
        print("\n  Exporting to Excel...")
        try:
            path = export_excel()
            print(f"  Saved to: {path}")
        except ImportError as e:
            print(f"  {e}")
            print("  Tip: use --csv instead (no extra packages needed)")

    summary = get_price_summary(days=DISPLAY_DAYS)
    if summary:
        print(f"\n  {'Product':<42} {'Retailer':<14} {'Latest':>8} {'MSRP':>7} {'Min':>7} {'Max':>7}")
        print("  " + "-" * 85)
        for row in summary[:15]:
            latest = row["latest_price_str"] or "N/A"
            msrp   = f"${row['msrp']:.2f}"    if row["msrp"]      else "N/A"
            lo     = f"${row['min_price']:.2f}" if row["min_price"] else "N/A"
            hi     = f"${row['max_price']:.2f}" if row["max_price"] else "N/A"
            print(
                f"  {row['name'][:41]:<42} {row['retailer']:<14} "
                f"{latest:>8} {msrp:>7} {lo:>7} {hi:>7}"
            )
        if len(summary) > 15:
            print(f"  ... and {len(summary) - 15} more products")
    else:
        print("\n  No data yet -- prices will appear after the first hourly log.")
        print("  Run with --log to force an immediate log.")

    print("\n" + "=" * 55 + "\n")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler()],
    )
    try:
        from tracker import CONFIG, PRODUCTS
        run_diagnostics(CONFIG, PRODUCTS)
    except ImportError:
        log.error("Run from tcg_tracker/ directory: python plugins/price_history.py")
